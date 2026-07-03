import os
import re
import shutil
from fastapi import FastAPI, HTTPException, Body, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import Optional, List
import openpyxl

from backend.rag_engine import get_engine, ask_deepseek_llm, compute_grounding, MERMAID_RULES, MERMAID_CLASSDEFS, RESOURCES_DIR

app = FastAPI(title="Integration Architect AI API")

# Enable CORS for frontend development
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

class ChatMessage(BaseModel):
    role: str
    content: str

class ChatRequest(BaseModel):
    query: str
    apiKey: Optional[str] = None
    history: Optional[List[ChatMessage]] = None

class ErrorDebugRequest(BaseModel):
    log: str
    apiKey: Optional[str] = None

class DiagramRepair(BaseModel):
    code: str
    error: str

class DiagramRequest(BaseModel):
    description: str
    diagramType: str = "flowchart"
    apiKey: Optional[str] = None
    repair: Optional[DiagramRepair] = None

class DDDAuditRequest(BaseModel):
    apiName: str
    endpoints: List[str]
    description: str
    systemsConnected: List[str]
    apiKey: Optional[str] = None

@app.on_event("startup")
def startup_event():
    # Warm up document search engine
    get_engine()

@app.on_event("shutdown")
def shutdown_event():
    pass

@app.get("/api/status")
def get_status():
    engine = get_engine()
    return {
        "status": "healthy",
        "chunks_loaded": len(engine.chunks),
        "resources_dir": RESOURCES_DIR,
        "indexed_files": list(set([c.filename for c in engine.chunks]))
    }

@app.post("/api/reload")
def reload_documents():
    engine = get_engine(force_reload=True)
    return {
        "status": "reloaded",
        "chunks_loaded": len(engine.chunks),
        "indexed_files": list(set([c.filename for c in engine.chunks]))
    }

@app.get("/api/documents")
def list_documents():
    engine = get_engine()
    docs = {}
    for chunk in engine.chunks:
        fname = chunk.filename
        if fname not in docs:
            docs[fname] = {
                "filename": fname,
                "chunks": 0,
                "type": chunk.metadata.get("type", "unknown")
            }
        docs[fname]["chunks"] += 1
    return list(docs.values())

@app.post("/api/documents/upload")
async def upload_document(file: UploadFile = File(...)):
    filename = file.filename
    if not filename:
        raise HTTPException(status_code=400, detail="No file uploaded")
    
    target_path = os.path.join(RESOURCES_DIR, filename)
    try:
        with open(target_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        
        # Trigger reload of RAG engine
        engine = get_engine(force_reload=True)
        return {
            "status": "success",
            "message": f"Successfully uploaded and indexed {filename}",
            "chunks_loaded": len(engine.chunks)
        }
    except Exception as e:
        if os.path.exists(target_path):
            os.remove(target_path)
        raise HTTPException(status_code=500, detail=f"Error saving file: {str(e)}")

@app.delete("/api/documents/delete/{filename}")
def delete_document(filename: str):
    if ".." in filename or "/" in filename or "\\" in filename:
        raise HTTPException(status_code=400, detail="Invalid filename format.")
        
    target_path = os.path.join(RESOURCES_DIR, filename)
    if not os.path.exists(target_path):
        raise HTTPException(status_code=404, detail=f"File {filename} not found.")
        
    try:
        os.remove(target_path)
        
        # Trigger reload of RAG engine
        engine = get_engine(force_reload=True)
        return {
            "status": "success",
            "message": f"Successfully deleted and re-indexed {filename}",
            "chunks_loaded": len(engine.chunks)
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error deleting file: {str(e)}")

@app.get("/api/search")
def search_chunks(q: str, limit: int = 5):
    if not q:
        raise HTTPException(status_code=400, detail="Query string 'q' is required.")
    engine = get_engine()
    results = engine.search(q, top_k=limit)
    return [
        {
            "id": c.id,
            "filename": c.filename,
            "content": c.content,
            "metadata": c.metadata
        }
        for c in results
    ]

@app.post("/api/chat")
async def chat_endpoint(request: ChatRequest):
    engine = get_engine()
    # Search document chunks related to user query
    chunks = engine.search(request.query, top_k=7)
    
    # Send to DeepSeek LLM with the last few conversation turns
    history = [{"role": m.role, "content": m.content} for m in (request.history or [])]
    response = await ask_deepseek_llm(request.query, chunks, request.apiKey, history=history)

    return {
        "response": response,
        "grounding": compute_grounding(response, chunks),
        "sources": [
            {
                "filename": c.filename,
                "content": c.content[:200] + "...",
                "metadata": c.metadata
            }
            for c in chunks
        ]
    }

@app.get("/api/mapping-visualizer")
def get_mapping_visualizer():
    # Find mapping Excel file
    excel_filename = "Enterprise_API_Directory_and_Data_Mappings.xlsx"
    filepath = os.path.join(RESOURCES_DIR, excel_filename)
    
    if not os.path.exists(filepath):
        raise HTTPException(status_code=404, detail="Enterprise mappings spreadsheet not found.")
        
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        sheets_data = {}
        
        for sheetname in wb.sheetnames:
            sheet = wb[sheetname]
            rows = []
            for row in sheet.iter_rows(values_only=True):
                # Filter out completely empty rows
                if any(c is not None for c in row):
                    rows.append([str(c) if c is not None else "" for c in row])
            sheets_data[sheetname] = rows
            
        return {
            "filename": excel_filename,
            "sheets": list(sheets_data.keys()),
            "data": sheets_data
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error reading mappings excel: {str(e)}")

@app.post("/api/debug-error")
async def debug_error_endpoint(request: ErrorDebugRequest):
    engine = get_engine()
    # Search specifically for error handling standard documents
    error_query = f"Error log error handler payload: {request.log}"
    chunks = engine.search(error_query, top_k=6)
    
    system_prompt = (
        "You are 'Integration Architect AI Error Log Debugger'. Paste and examine the user's raw error log.\n"
        "1. Identify the Mule Error Type (e.g. VALIDATION:INVALID_BOOLEAN, HTTP:CONNECTIVITY, etc.) or description.\n"
        "2. Retrieve details from 'Sample error test 3.pdf' or 'MuleSoft_Development_Best_Practices_and_Standards.docx' "
        "to classify the error category:\n"
        "   - Business / Data Error: Bad input data. Action: ACK message, log payload/message to Salesforce Custom Object Integration_Error__c.\n"
        "   - System / Transient Error: Infrastructure/network down. Action: NACK/Rollback transaction, retry (max 3 times), DLQ on exhaustion.\n"
        "3. Provide the standard operational recommendation (Business remediation vs Platform recovery).\n"
        "4. Output the recommended DataWeave Mapping or Mule XML Config block (such as VM/JMS listener redelivery-policy, or error-handler structure) "
        "relevant to this error.\n"
        "Answer in clean Markdown format."
    )
    
    # Combine chunks
    context_str = "\n\n---\n\n".join([chunk.content for chunk in chunks])
    full_prompt = (
        f"CONTEXT FROM REPOSITORY:\n{context_str}\n\n"
        f"USER ERROR LOG:\n{request.log}\n\n"
        "Provide a complete analysis following the instructions."
    )
    
    # We call ask_deepseek_llm indirectly by constructing a dummy chunk list
    # and custom query to match DeepSeek endpoint requirements
    from backend.rag_engine import DocumentChunk
    dummy_chunk = DocumentChunk("error_analysis", "System context", system_prompt)
    response = await ask_deepseek_llm(full_prompt, [dummy_chunk], request.apiKey)
    
    return {
        "analysis": response,
        "grounding": compute_grounding(response, chunks),
        "relevant_sources": [c.filename for c in chunks]
    }

@app.post("/api/generate-diagram")
async def generate_diagram_endpoint(request: DiagramRequest):
    engine = get_engine()
    # Reference diagrams (OCR-indexed images) first, then standard context; dedupe by id
    diagram_chunks = engine.search(request.description, top_k=4, where={"type": "diagram"})
    text_chunks = engine.search(f"{request.description} API integration architecture flow systems", top_k=8)
    seen = {c.id for c in diagram_chunks}
    chunks = diagram_chunks + [c for c in text_chunks if c.id not in seen]

    if request.repair:
        query = (
            "The following Mermaid diagram failed to render in the browser with a parse error. "
            "Fix the syntax and return the corrected diagram.\n\n"
            f"MERMAID CODE:\n```mermaid\n{request.repair.code}\n```\n\n"
            f"PARSE ERROR:\n{request.repair.error}\n\n"
            f"{MERMAID_RULES}"
        )
    else:
        query = (
            f"Create a professional '{request.diagramType}' architecture diagram in Mermaid for the following scenario, "
            f"grounded in the enterprise integration documentation context:\n\n{request.description}\n\n{MERMAID_RULES}"
        )

    response = await ask_deepseek_llm(query, chunks, request.apiKey)

    match = re.search(r"```mermaid\s*\n([\s\S]*?)```", response)
    mermaid_code = match.group(1).strip() if match else None
    notes = (response[:match.start()] + response[match.end():]).strip() if match else None

    # LLMs regularly drop the classDef block; inject it so dark-theme styling is guaranteed
    if mermaid_code and mermaid_code.lstrip().startswith(("flowchart", "---")) and "classDef" not in mermaid_code:
        mermaid_code += "\n" + MERMAID_CLASSDEFS

    return {
        "mermaid": mermaid_code,
        "notes": notes,
        "raw": None if mermaid_code else response,
        "grounding": compute_grounding(response, chunks),
        "sources": [
            {
                "filename": c.filename,
                "content": c.content[:200] + "...",
                "metadata": c.metadata
            }
            for c in chunks
        ]
    }

@app.post("/api/audit-ddd")
async def audit_ddd_endpoint(request: DDDAuditRequest):
    engine = get_engine()
    # Search for DDD compliance files
    search_query = f"Domain Driven Design DDD rules naming convention systems canonical schema"
    chunks = engine.search(search_query, top_k=6)
    
    prompt = (
        "Review this proposed API design against the enterprise Domain-Driven Design (DDD) specifications "
        "laid out in 'Sample Mulesoft docs1.pdf':\n\n"
        f"API Name: {request.apiName}\n"
        f"Description: {request.description}\n"
        f"Endpoints: {', '.join(request.endpoints)}\n"
        f"Backend Systems Connected: {', '.join(request.systemsConnected)}\n\n"
        "Audit the proposal on:\n"
        "1. Naming Conventions: (Is the API noun-based/domain-centric? Are regional identifiers na/emea or system names sf/sap leaked in the API name?)\n"
        "2. System Decoupling: (Does it connect directly to System APIs, or bypass process layers?)\n"
        "3. Canonical Models: (Does it represent a unified business domain case/shipment/order instead of system-specific views?)\n"
        "4. Queueing/Ingestion Split: (Should event ingestion be separated from processing via listener and publisher flows?)\n\n"
        "Provide a compliance score (0-100%), list of violations, and recommended architecture changes in Markdown."
    )
    
    response = await ask_deepseek_llm(prompt, chunks, request.apiKey)
    return {
        "audit": response,
        "grounding": compute_grounding(response, chunks)
    }
